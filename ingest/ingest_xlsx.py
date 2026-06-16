#!/usr/bin/env python3
"""
ingest_xlsx.py — Ingest EU Tariff Portal Excel exports into the AImport DB.

Handles the 5-file export set:
  Nomenclature EN.xlsx, Declarable codes.xlsx, Duties Import 01-99.xlsx,
  Footnotes descriptions.xlsx, Geographical areas descriptions.xlsx

DESIGN RULES (same as XML path):
  - Verbatim duty text is ALWAYS stored (measure.duty_raw) — parsing adds
    structure, never replaces the legal source text.
  - Duty forms we cannot compute (Meursing EA/ADSZ agri-components) are
    FLAGGED, not guessed.
  - Unparsed rows are counted and reported; ingest declares itself
    incomplete if counts exceed tolerance.
KNOWN LIMITATION (v1): this export set contains no geographical-group
  membership table. Origin queries therefore match: exact country code +
  ERGA OMNES (1011). Group-targeted measures (e.g. "GSP countries") are
  stored and shown by name, but not auto-resolved to member countries yet.

USAGE:
    python3 ingest/ingest_xlsx.py <folder_with_xlsx> <database.sqlite>
"""
import hashlib
import re
import sqlite3
import sys
from datetime import datetime, timezone, date
from pathlib import Path
from openpyxl import load_workbook

SCHEMA = Path(__file__).resolve().parent.parent / "db" / "schema.sql"


# ---------------------------------------------------------------- helpers
def parse_date(v):
    """Normalise the two date dialects in these files to ISO YYYY-MM-DD."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime) or isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    raise ValueError(f"unrecognised date: {v!r}")


def split_code(v):
    """'0101210000 10' -> ('0101210000', '10'); '0100000000' -> (code,'80')."""
    s = str(v).strip()
    m = re.fullmatch(r"(\d{10})(?:\s+(\d{2}))?", s)
    if not m:
        raise ValueError(f"bad goods code: {v!r}")
    return m.group(1), m.group(2) or "80"


# ----- duty text grammar (designed against the 205 observed patterns) ----
RE_ADVAL = re.compile(r"(\d+(?:\.\d+)?)\s*%")
RE_SPECIFIC = re.compile(r"(\d+(?:\.\d+)?)\s*EUR\s*/?\s*([A-Z]{2,4}(?:\s?[A-Z])?)")
RE_CERT = re.compile(r"cert:\s*([A-Z])-(\w{3})")
RE_COND_SEG = re.compile(r"(?:^|;)\s*([A-Z])\s*(?:cert:\s*[A-Z]-\w{3}\s*)?(?:[\d,]+(?:\.\d+)?\s*(?:EUR/)?\w+\s*)?\((\d{2})\)")
MEURSING_MARKERS = ("EA", "ADSZ", "ADFM")


def parse_duty(text):
    """Returns (components, conditions, flags).
    components: list of (duty_expression_id, amount, monetary_unit, unit)
       expression ids: '01' ad-valorem, '04' specific(+), 'MEURSING', 'SUPP'
    conditions: list of (condition_code, cert_type, cert_code, action_code)
    flags: set of strings for human attention."""
    comps, conds, flags = [], [], set()
    if text is None:
        return comps, conds, flags
    s = str(text).strip()
    if not s:
        return comps, conds, flags

    # supplementary-unit-only rows like 'NAR'
    if re.fullmatch(r"[A-Z]{3}(\s+[A-Z])?", s) and s != "NIH":
        comps.append(("SUPP", None, None, s.split()[0]))
        return comps, conds, flags

    # 'NIHIL' = legally nil duty
    if s.upper().startswith("NIHIL"):
        comps.append(("01", 0.0, None, None))
        flags.add("NIHIL")
        return comps, conds, flags

    if any(m in s.split() or f"+{m}" in s.replace(" ", "") for m in MEURSING_MARKERS) \
            or re.search(r"\bEA\b|\bADSZ\b|\bADFM\b", s):
        flags.add("MEURSING")  # agri-component duty: needs composition data

    is_conditional = s.startswith("Cond:")
    if is_conditional:
        flags.add("CONDITIONAL")
        body = s[5:]  # strip 'Cond:' so the first segment is scanned too
        for m in RE_COND_SEG.finditer(body):
            letter, action = m.group(1), m.group(2)
            seg_start = m.start()
            seg_end = body.find(";", m.end())
            seg = body[seg_start: seg_end if seg_end != -1 else len(body)]
            certm = RE_CERT.search(seg)
            conds.append((letter,
                          certm.group(1) if certm else None,
                          certm.group(2) if certm else None,
                          action))
        # do NOT turn conditional sub-duties into flat components: the duty
        # depends on which condition is met — engine must show alternatives.
        return comps, conds, flags

    # plain duty: ad valorem and/or specific parts
    for m in RE_ADVAL.finditer(s):
        comps.append(("01", float(m.group(1)), None, None))
    for m in RE_SPECIFIC.finditer(s):
        comps.append(("04", float(m.group(1)), "EUR", m.group(2).strip()))
    # 'N EUR DTN' (no slash) form:
    if not RE_SPECIFIC.search(s):
        for m in re.finditer(r"(\d+(?:\.\d+)?)\s*EUR\s+([A-Z]{3})", s):
            comps.append(("04", float(m.group(1)), "EUR", m.group(2)))
    if not comps and "MEURSING" not in flags:
        flags.add("UNPARSED")
    return comps, conds, flags


# ---------------------------------------------------------------- ingest
def file_sha(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def rows_of(path, sheet=None):
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it)
    for row in it:
        if any(c is not None for c in row):
            yield row
    wb.close()


def ingest(folder: Path, db_path: Path):
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA.read_text())
    stats = {"loaded": 0, "errors": [], "meursing": 0, "unparsed_duty": 0,
             "conditional": 0}

    def log_snapshot(fname):
        cur = conn.execute(
            "INSERT INTO ingest_log (source_file, file_sha256, loaded_at, records_loaded) VALUES (?,?,?,0)",
            (fname, file_sha(folder / fname),
             datetime.now(timezone.utc).isoformat(timespec="seconds")))
        return cur.lastrowid

    # ---- 1. Nomenclature EN ------------------------------------------
    f = "Nomenclature EN.xlsx"
    snap = log_snapshot(f)
    n = 0
    seen_sids = set()
    for row in rows_of(folder / f):
        try:
            code_raw, start, end, lang, hier, indent, desc, dstart = row[:8]
            code, suffix = split_code(code_raw)
            sid = int(code + suffix)          # synthetic stable key
            if sid not in seen_sids:
                seen_sids.add(sid)
                conn.execute("INSERT INTO goods_nomenclature VALUES (?,?,?,?,?,?,?,?)",
                             (sid, code, suffix, parse_date(start), parse_date(end),
                              0, 0, snap))
                level = (str(indent).count("-") if indent else 0)
                conn.execute("INSERT INTO goods_nomenclature_indent VALUES (?,?,?,?,?)",
                             (sid, level, parse_date(start) or "1900-01-01", parse_date(end), snap))
            conn.execute("INSERT INTO goods_nomenclature_description VALUES (?,?,?,?,?,?)",
                         (sid, lang or "EN", str(desc or "").strip(),
                          parse_date(dstart) or parse_date(start) or "1900-01-01",
                          parse_date(end), snap))
            n += 1
        except Exception as e:
            stats["errors"].append(f"[nomenclature] {e} :: {str(row)[:90]}")
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, snap))
    stats["loaded"] += n

    # ---- 2. Declarable codes (IS_LEAF) -------------------------------
    f = "Declarable codes.xlsx"
    snap = log_snapshot(f)
    n = 0
    for row in rows_of(folder / f):
        try:
            code_raw, start, decl_start, is_leaf, end = row[:5]
            code, suffix = split_code(code_raw)
            sid = int(code + suffix)
            conn.execute("UPDATE goods_nomenclature SET is_leaf=? WHERE sid=?",
                         (int(is_leaf or 0), sid))
            n += 1
        except Exception as e:
            stats["errors"].append(f"[declarable] {e} :: {str(row)[:90]}")
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, snap))
    stats["loaded"] += n

    # ---- 3. Duties Import --------------------------------------------
    f = "Duties Import 01-99.xlsx"
    snap = log_snapshot(f)
    n = 0
    msid = 0
    for row in rows_of(folder / f):
        try:
            (code_raw, add_code, order_no, start, end, red, origin_name,
             mtype_name, legal, duty, origin_code, mtype_code) = row[:12]
            code, _ = split_code(code_raw)
            msid += 1
            comps, conds, flags = parse_duty(duty)
            conn.execute("INSERT INTO measure VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                         (msid, code, None, str(mtype_code or "").strip(),
                          str(origin_code or "").strip(), None,
                          str(add_code).strip() if add_code else None,
                          str(legal or "").strip(),
                          str(duty).strip() if duty is not None else None,
                          str(origin_name or "").strip(),
                          parse_date(start), parse_date(end), snap))
            conn.execute("INSERT OR IGNORE INTO measure_type VALUES (?,?,?)",
                         (str(mtype_code or "").strip(), str(mtype_name or "").strip(), snap))
            for de, amt, mu, unit in comps:
                conn.execute("INSERT INTO measure_component VALUES (?,?,?,?,?,?)",
                             (msid, de, amt, mu, unit, snap))
            for cc, ct, ccode, action in conds:
                conn.execute("INSERT INTO measure_condition VALUES (?,?,?,?,?,?)",
                             (msid, cc, ct, ccode, action, snap))
            if "MEURSING" in flags:
                stats["meursing"] += 1
            if "UNPARSED" in flags:
                stats["unparsed_duty"] += 1
            if "CONDITIONAL" in flags:
                stats["conditional"] += 1
            n += 1
        except Exception as e:
            stats["errors"].append(f"[duties] {e} :: {str(row)[:90]}")
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, snap))
    stats["loaded"] += n

    # ---- 4. Footnotes (EN only) ---------------------------------------
    f = "Footnotes descriptions.xlsx"
    snap = log_snapshot(f)
    n = 0
    for row in rows_of(folder / f):
        try:
            fid, lang, desc, start, dstart, end = row[:6]
            if (lang or "").strip() != "EN":
                continue
            fid = str(fid).strip()
            conn.execute("INSERT INTO footnote VALUES (?,?,?,?)",
                         (fid[:2], fid[2:], str(desc or "").strip(), snap))
            n += 1
        except Exception as e:
            stats["errors"].append(f"[footnotes] {e} :: {str(row)[:90]}")
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, snap))
    stats["loaded"] += n

    # ---- 5. Geographical areas (EN only) -------------------------------
    f = "Geographical areas descriptions.xlsx"
    snap = log_snapshot(f)
    n = 0
    gsid = 0
    seen = set()
    for row in rows_of(folder / f):
        try:
            area_id, lang, abbrev, desc, start, dstart, end = row[:7]
            if (lang or "").strip() != "EN":
                continue
            area_id = str(area_id).strip()
            if area_id in seen:
                continue
            seen.add(area_id)
            gsid += 1
            is_group = 1 if area_id.isdigit() else 0
            conn.execute("INSERT INTO geographical_area VALUES (?,?,?,?,?,?)",
                         (gsid, area_id, is_group, parse_date(start) or "1900-01-01",
                          parse_date(end), snap))
            conn.execute("INSERT INTO geographical_area_description VALUES (?,?,?,?)",
                         (gsid, "EN", str(desc or "").strip(), snap))
            n += 1
        except Exception as e:
            stats["errors"].append(f"[geo] {e} :: {str(row)[:90]}")
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, snap))
    stats["loaded"] += n

    conn.commit()
    conn.close()
    return stats


def main():
    folder, db = Path(sys.argv[1]), Path(sys.argv[2])
    stats = ingest(folder, db)
    print(f"Loaded rows:          {stats['loaded']}")
    print(f"Conditional duties:   {stats['conditional']} (conditions extracted; duty alternatives kept in raw text)")
    print(f"Meursing/agri duties: {stats['meursing']} (FLAGGED — need composition data, never auto-computed)")
    print(f"Unparsed duty texts:  {stats['unparsed_duty']} (raw text preserved)")
    if stats["errors"]:
        print(f"\n!! ROW ERRORS ({len(stats['errors'])}) — first 10:")
        for e in stats["errors"][:10]:
            print("  ", e)
    total = stats["loaded"]
    bad = stats["unparsed_duty"] + len(stats["errors"])
    if bad > total * 0.005:
        print(f"\nRESULT: INGEST INCOMPLETE ({bad} problems > 0.5% tolerance) — "
              "do not use for classification.")
        sys.exit(2)
    print("\nRESULT: CLEAN INGEST (within tolerance).")


if __name__ == "__main__":
    main()
