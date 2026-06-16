#!/usr/bin/env python3
"""
ingest_xlsx_v2.py — Ingest the four supplementary EU Tariff Portal sheets:
  Measure conditions, Measure exclusions, Legal basis, Additional codes.

These drive AImport's compliance OUTPUT (certificates, licences, exclusions,
legal citations), so correctness here is treated as non-negotiable:

  - VERBATIM PRESERVATION. Raw certificate codes, action codes and dates are
    stored exactly; interpretation is a presentation-layer concern, never a
    mutation of the source.
  - COMPOSITE-KEY JOIN. Conditions/exclusions link to `measure` by
    (goods_code, add_code, origin_code, measure_type, validity_start) — the
    same business key the duties sheet uses.
  - EXCLUSIONS ARE SAFETY-CRITICAL. A group measure (e.g. erga-omnes control)
    that carves out a country must NOT be applied to that country. Storing
    exclusions prevents false-positive compliance flags.

Run AFTER ingest_xlsx.py, against the same database:
    python3 ingest/ingest_xlsx_v2.py <folder> <database.sqlite>
"""
import hashlib
import re
import sys
from datetime import datetime, timezone, date
from pathlib import Path
from openpyxl import load_workbook

SCHEMA = Path(__file__).resolve().parent.parent / "db" / "schema.sql"


def parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    raise ValueError(f"bad date {v!r}")


def norm_code(v):
    if v is None:
        return None
    return re.sub(r"\s+", "", str(v).strip())


def file_sha(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for c in iter(lambda: f.read(1 << 20), b""):
            h.update(c)
    return h.hexdigest()


def rows_of(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else h for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    for row in it:
        if any(c is not None for c in row):
            yield row, idx
    wb.close()


def ingest(folder: Path, conn):
    conn.executescript(SCHEMA.read_text())
    stats = {}

    def snap(fname):
        cur = conn.execute(
            "INSERT INTO ingest_log (source_file, file_sha256, loaded_at, records_loaded) VALUES (?,?,?,0)",
            (fname, file_sha(folder / fname),
             datetime.now(timezone.utc).isoformat(timespec="seconds")))
        return cur.lastrowid

    # ---- Legal basis ---------------------------------------------------
    f = "Legal basis.xlsx"
    s = snap(f)
    n = 0
    for row, I in rows_of(folder / f):
        conn.execute("INSERT OR REPLACE INTO legal_basis VALUES (?,?,?,?,?)",
                     (str(row[I['Legal base']]).strip(),
                      row[I['Off. Journal']], str(row[I['Page']]) if row[I['Page']] is not None else None,
                      parse_date(row[I['Publ. date']]), s))
        n += 1
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, s))
    stats["legal_basis"] = n

    # ---- Additional codes (all languages; EN prioritised at query) -----
    f = "Additional codes descriptions.xlsx"
    s = snap(f)
    n = 0
    for row, I in rows_of(folder / f):
        if (row[I['Language']] or "").strip() != "EN":
            continue
        conn.execute("INSERT INTO additional_code_description VALUES (?,?,?,?)",
                     (norm_code(row[I['Add code']]), "EN",
                      str(row[I['Description']] or "").strip(), s))
        n += 1
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, s))
    stats["additional_codes_EN"] = n

    # ---- Measure conditions -------------------------------------------
    f = "Measure conditions.xlsx"
    s = snap(f)
    n = 0
    for row, I in rows_of(folder / f):
        conn.execute("INSERT INTO measure_condition_v2 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                     (norm_code(row[I['Goods code']]),
                      norm_code(row[I['Add code']]),
                      str(row[I['Origin code']] or "").strip(),
                      str(row[I['Meas. type code']] or "").strip(),
                      parse_date(row[I['Start date']]),
                      parse_date(row[I['End date']]),
                      row[I['Meas. cond']],
                      int(row[I['Sequence']]) if row[I['Sequence']] is not None else None,
                      norm_code(row[I['Certificate']]),
                      float(row[I['Cond. amount']]) if row[I['Cond. amount']] is not None else None,
                      row[I['Mon. unit']], row[I['Meas. unit']],
                      str(row[I['Meas. action']]) if row[I['Meas. action']] is not None else None,
                      s))
        n += 1
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, s))
    stats["conditions"] = n

    # ---- Measure exclusions -------------------------------------------
    f = "Measure exclusions.xlsx"
    s = snap(f)
    n = 0
    for row, I in rows_of(folder / f):
        conn.execute("INSERT INTO measure_exclusion VALUES (?,?,?,?,?,?,?,?)",
                     (norm_code(row[I['Goods code']]),
                      norm_code(row[I['Add code']]),
                      str(row[I['Origin code']] or "").strip(),
                      str(row[I['Meas. type code']] or "").strip(),
                      parse_date(row[I['Start date']]),
                      parse_date(row[I['End date']]),
                      str(row[I['Excluded country code']] or "").strip(), s))
        n += 1
    conn.execute("UPDATE ingest_log SET records_loaded=? WHERE snapshot_id=?", (n, s))
    stats["exclusions"] = n

    conn.commit()
    return stats


def main():
    import sqlite3
    folder, db = Path(sys.argv[1]), Path(sys.argv[2])
    conn = sqlite3.connect(db)
    stats = ingest(folder, conn)
    for k, v in stats.items():
        print(f"  {k:<22}{v:>9}")
    print("\nRESULT: supplementary sheets ingested.")


if __name__ == "__main__":
    main()
